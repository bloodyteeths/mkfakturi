#!/usr/bin/env python3
"""
Smart CRM Excel Generator for Facturino
Creates a comprehensive CRM spreadsheet for Google Sheets
"""

import csv
import os
from datetime import datetime, timedelta
import random

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
    from openpyxl.chart import PieChart, BarChart, Reference
    from openpyxl.worksheet.datavalidation import DataValidation
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("openpyxl not installed. Installing...")
    import subprocess
    subprocess.check_call(['pip3', 'install', 'openpyxl'])
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
    from openpyxl.chart import PieChart, BarChart, Reference
    from openpyxl.worksheet.datavalidation import DataValidation

# Paths
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
LEADS_DIR = os.path.join(BASE_DIR, 'storage', 'leads')
OUTPUT_FILE = os.path.join(BASE_DIR, 'Facturino_Smart_CRM.xlsx')

# Styling
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
ALT_ROW_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Status colors
STATUS_COLORS = {
    "Нов": "70AD47",      # Green
    "Контактиран": "FFC000",  # Yellow
    "Заинтересиран": "5B9BD5",  # Blue
    "Демо закажано": "7030A0",  # Purple
    "Преговори": "ED7D31",  # Orange
    "Добиен": "00B050",  # Dark Green
    "Изгубен": "C00000",  # Red
    "Неактивен": "808080",  # Gray
}

def load_leads():
    """Load leads from CSV files"""
    leads = []

    # Load ISOS leads
    isos_path = os.path.join(LEADS_DIR, 'isos_leads.csv')
    if os.path.exists(isos_path):
        with open(isos_path, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                row['source'] = 'ISOS'
                row['type'] = 'Сметководствена компанија'
                leads.append(row)

    # Load smetkovoditeli leads
    smk_path = os.path.join(LEADS_DIR, 'smetkovoditeli_leads.csv')
    if os.path.exists(smk_path):
        with open(smk_path, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                row['source'] = 'Сметководители.мк'
                row['type'] = 'Сметководствена компанија'
                leads.append(row)

    return leads

def create_workbook():
    """Create the main workbook with all sheets"""
    wb = Workbook()

    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    return wb

def create_dashboard(wb):
    """Create dashboard sheet with overview stats"""
    ws = wb.create_sheet("📊 Dashboard", 0)

    # Title
    ws['A1'] = "FACTURINO SMART CRM"
    ws['A1'].font = Font(size=24, bold=True, color="4472C4")
    ws.merge_cells('A1:F1')

    ws['A2'] = f"Последно ажурирање: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    ws['A2'].font = Font(italic=True, color="808080")

    # Summary cards
    ws['A4'] = "📈 ПРЕГЛЕД"
    ws['A4'].font = Font(size=16, bold=True)

    # Stats formulas (will reference Companies sheet)
    stats = [
        ("Вкупно компании", "=COUNTA(Компании!A:A)-1"),
        ("Нови лидови", '=COUNTIF(Компании!G:G,"Нов")'),
        ("Контактирани", '=COUNTIF(Компании!G:G,"Контактиран")'),
        ("Заинтересирани", '=COUNTIF(Компании!G:G,"Заинтересиран")'),
        ("Демо закажано", '=COUNTIF(Компании!G:G,"Демо закажано")'),
        ("Добиени клиенти", '=COUNTIF(Компании!G:G,"Добиен")'),
        ("Изгубени", '=COUNTIF(Компании!G:G,"Изгубен")'),
    ]

    for i, (label, formula) in enumerate(stats):
        row = 6 + i
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = formula
        ws[f'B{row}'].font = Font(size=14, bold=True, color="4472C4")

    # Conversion rate
    ws['A14'] = "Стапка на конверзија"
    ws['A14'].font = Font(bold=True)
    ws['B14'] = '=IF(B6>0,B11/B6*100,0)'
    ws['C14'] = "%"

    # Pipeline value
    ws['A16'] = "💰 ВРЕДНОСТ НА PIPELINE"
    ws['A16'].font = Font(size=16, bold=True)

    ws['A17'] = "Потенцијална вредност"
    ws['B17'] = "=SUMIF(Компании!G:G,\"<>Изгубен\",Компании!I:I)"
    ws['B17'].number_format = '#,##0 "МКД"'

    ws['A18'] = "Очекувана вредност (weighted)"
    ws['B18'] = "=SUMPRODUCT((Компании!G:G=\"Заинтересиран\")*Компании!I:I*0.3)+SUMPRODUCT((Компании!G:G=\"Демо закажано\")*Компании!I:I*0.5)+SUMPRODUCT((Компании!G:G=\"Преговори\")*Компании!I:I*0.7)+SUMPRODUCT((Компании!G:G=\"Добиен\")*Компании!I:I*1)"
    ws['B18'].number_format = '#,##0 "МКД"'

    # Tasks due
    ws['A20'] = "📋 ЗАДАЧИ"
    ws['A20'].font = Font(size=16, bold=True)

    ws['A21'] = "Задачи за денес"
    ws['B21'] = '=COUNTIF(Активности!D:D,TODAY())'

    ws['A22'] = "Задачи оваа недела"
    ws['B22'] = '=COUNTIFS(Активности!D:D,">="&TODAY(),Активности!D:D,"<="&TODAY()+7)'

    ws['A23'] = "Закаснети задачи"
    ws['B23'] = '=COUNTIFS(Активности!D:D,"<"&TODAY(),Активности!E:E,"<>Завршено")'
    ws['B23'].font = Font(color="C00000", bold=True)

    # Column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15

    return ws

def create_companies_sheet(wb, leads):
    """Create companies sheet with all leads"""
    ws = wb.create_sheet("Компании", 1)

    # Headers
    headers = [
        "ID", "Име на компанија", "Email", "Телефон", "Град",
        "Извор", "Статус", "Приоритет", "Очекувана вредност (МКД)",
        "Датум додадено", "Последен контакт", "Следен чекор",
        "Одговорно лице", "Веб страна", "Забелешки"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    # Add leads data
    for i, lead in enumerate(leads, 2):
        ws.cell(row=i, column=1, value=i-1)  # ID
        ws.cell(row=i, column=2, value=lead.get('company_name', ''))
        ws.cell(row=i, column=3, value=lead.get('email', ''))
        ws.cell(row=i, column=4, value=lead.get('phone', ''))
        ws.cell(row=i, column=5, value=lead.get('city', ''))
        ws.cell(row=i, column=6, value=lead.get('source', ''))
        ws.cell(row=i, column=7, value='Нов')  # Default status
        ws.cell(row=i, column=8, value='Среден')  # Default priority
        ws.cell(row=i, column=9, value=0)  # Expected value
        ws.cell(row=i, column=10, value=datetime.now().strftime('%d.%m.%Y'))
        ws.cell(row=i, column=14, value=lead.get('website', ''))

        # Alternate row coloring
        if i % 2 == 0:
            for col in range(1, len(headers) + 1):
                ws.cell(row=i, column=col).fill = ALT_ROW_FILL

    # Data validation for Status
    status_dv = DataValidation(
        type="list",
        formula1='"Нов,Контактиран,Заинтересиран,Демо закажано,Преговори,Добиен,Изгубен,Неактивен"',
        allow_blank=True
    )
    status_dv.error = "Изберете валиден статус"
    status_dv.errorTitle = "Невалиден статус"
    ws.add_data_validation(status_dv)
    status_dv.add(f'G2:G{len(leads)+100}')

    # Data validation for Priority
    priority_dv = DataValidation(
        type="list",
        formula1='"Висок,Среден,Низок"',
        allow_blank=True
    )
    ws.add_data_validation(priority_dv)
    priority_dv.add(f'H2:H{len(leads)+100}')

    # Conditional formatting for status
    from openpyxl.formatting.rule import FormulaRule

    # Green for "Добиен"
    ws.conditional_formatting.add(f'G2:G{len(leads)+100}',
        FormulaRule(formula=['$G2="Добиен"'],
                   fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")))

    # Red for "Изгубен"
    ws.conditional_formatting.add(f'G2:G{len(leads)+100}',
        FormulaRule(formula=['$G2="Изгубен"'],
                   fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")))

    # Yellow for "Заинтересиран"
    ws.conditional_formatting.add(f'G2:G{len(leads)+100}',
        FormulaRule(formula=['$G2="Заинтересиран"'],
                   fill=PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")))

    # Column widths
    column_widths = [6, 40, 30, 15, 15, 20, 15, 12, 18, 12, 12, 20, 15, 25, 30]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64+i) if i <= 26 else 'A' + chr(64+i-26)].width = width

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Auto-filter
    ws.auto_filter.ref = f"A1:O{len(leads)+1}"

    return ws

def create_contacts_sheet(wb):
    """Create contacts sheet for individual people"""
    ws = wb.create_sheet("Контакти", 2)

    headers = [
        "ID", "Име и презиме", "Компанија", "Позиција", "Email",
        "Телефон", "LinkedIn", "Примарен контакт", "Датум додадено", "Забелешки"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER

    # Data validation for Primary Contact
    primary_dv = DataValidation(
        type="list",
        formula1='"Да,Не"',
        allow_blank=True
    )
    ws.add_data_validation(primary_dv)
    primary_dv.add('H2:H1000')

    # Column widths
    column_widths = [6, 25, 35, 20, 30, 15, 30, 15, 12, 40]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64+i)].width = width

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = "A1:J1"

    return ws

def create_activities_sheet(wb):
    """Create activities sheet for tracking calls, emails, meetings"""
    ws = wb.create_sheet("Активности", 3)

    headers = [
        "ID", "Компанија", "Тип", "Датум закажано", "Статус",
        "Опис", "Резултат", "Следен чекор", "Одговорно лице", "Датум креирано"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER

    # Data validation for Type
    type_dv = DataValidation(
        type="list",
        formula1='"Повик,Email,Состанок,Демо,Follow-up,Друго"',
        allow_blank=True
    )
    ws.add_data_validation(type_dv)
    type_dv.add('C2:C1000')

    # Data validation for Status
    status_dv = DataValidation(
        type="list",
        formula1='"Закажано,Во тек,Завршено,Откажано"',
        allow_blank=True
    )
    ws.add_data_validation(status_dv)
    status_dv.add('E2:E1000')

    # Conditional formatting for overdue tasks
    ws.conditional_formatting.add('D2:D1000',
        FormulaRule(formula=['AND($D2<TODAY(),$E2<>"Завршено",$D2<>"")'],
                   fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")))

    # Column widths
    column_widths = [6, 35, 12, 15, 12, 40, 30, 30, 15, 12]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64+i)].width = width

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = "A1:J1"

    return ws

def create_pipeline_sheet(wb):
    """Create sales pipeline/deals sheet"""
    ws = wb.create_sheet("Pipeline", 4)

    headers = [
        "ID", "Име на deal", "Компанија", "Вредност (МКД)", "Фаза",
        "Веројатност %", "Очекуван датум", "Производ/План", "Одговорно лице",
        "Датум креирано", "Забелешки"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER

    # Data validation for Phase
    phase_dv = DataValidation(
        type="list",
        formula1='"Квалификација,Демо,Понуда,Преговори,Затворено-Добиено,Затворено-Изгубено"',
        allow_blank=True
    )
    ws.add_data_validation(phase_dv)
    phase_dv.add('E2:E1000')

    # Data validation for Product/Plan
    product_dv = DataValidation(
        type="list",
        formula1='"Основен план,Про план,Премиум план,Enterprise"',
        allow_blank=True
    )
    ws.add_data_validation(product_dv)
    product_dv.add('H2:H1000')

    # Column widths
    column_widths = [6, 30, 35, 15, 15, 12, 15, 15, 15, 12, 40]
    for i, width in enumerate(column_widths, 1):
        col_letter = chr(64+i) if i <= 26 else 'A' + chr(64+i-26)
        ws.column_dimensions[col_letter].width = width

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = "A1:K1"

    return ws

def create_email_templates_sheet(wb):
    """Create email templates sheet"""
    ws = wb.create_sheet("Email Templates", 5)

    headers = ["Име на шаблон", "Предмет", "Содржина"]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER

    # Sample templates
    templates = [
        (
            "Прв контакт",
            "Дигитализирајте го вашето сметководство со Facturino",
            """Почитувани,

Ви се обраќам од Facturino - македонска платформа за е-фактурирање и сметководство.

Забележавме дека вашата компанија е регистрирана во ISOS регистарот и сметаме дека нашето решение може значително да ви го олесни секојдневното работење.

Facturino нуди:
• Автоматско креирање на е-фактури согласно МК регулатива
• Интеграција со банки за автоматско книжење
• Партнерска програма со атрактивни провизии

Дали би имале 15 минути за краток разговор?

Со почит,
[Вашето име]
Facturino тим"""
        ),
        (
            "Follow-up после демо",
            "Следни чекори после нашата демо презентација",
            """Почитувани [Име],

Ви благодарам за времето посветено на нашата демо презентација.

Како што разговаравме, Facturino може да ви помогне со:
[Персонализирајте според дискусијата]

Следни чекори:
1. Ќе ви испратам пристап до пробна верзија
2. Закажуваме follow-up повик за [датум]

Дали имате дополнителни прашања?

Со почит,
[Вашето име]"""
        ),
        (
            "Партнерска покана",
            "Ексклузивна партнерска програма за сметководители",
            """Почитувани,

Facturino воведува партнерска програма специјално дизајнирана за сметководствени бироа.

Придобивки:
• 20% провизија за секој препорачан клиент
• Бесплатна Про верзија за вашето биро
• Приоритетна техничка поддршка
• Ко-брендирање можности

Дали сакате да дознаете повеќе за партнерството?

Со почит,
[Вашето име]
Facturino Partner Program"""
        ),
    ]

    for i, (name, subject, content) in enumerate(templates, 2):
        ws.cell(row=i, column=1, value=name)
        ws.cell(row=i, column=2, value=subject)
        ws.cell(row=i, column=3, value=content)
        ws.cell(row=i, column=3).alignment = Alignment(wrap_text=True)
        ws.row_dimensions[i].height = 150

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 80

    return ws

def create_settings_sheet(wb):
    """Create settings sheet with dropdown options"""
    ws = wb.create_sheet("⚙️ Поставки", 6)

    ws['A1'] = "ПОСТАВКИ И ОПЦИИ"
    ws['A1'].font = Font(size=16, bold=True)

    # Status options
    ws['A3'] = "Статуси на компании:"
    ws['A3'].font = Font(bold=True)
    statuses = ["Нов", "Контактиран", "Заинтересиран", "Демо закажано", "Преговори", "Добиен", "Изгубен", "Неактивен"]
    for i, status in enumerate(statuses, 4):
        ws[f'A{i}'] = status

    # Priority options
    ws['C3'] = "Приоритети:"
    ws['C3'].font = Font(bold=True)
    priorities = ["Висок", "Среден", "Низок"]
    for i, priority in enumerate(priorities, 4):
        ws[f'C{i}'] = priority

    # Activity types
    ws['E3'] = "Типови активности:"
    ws['E3'].font = Font(bold=True)
    activities = ["Повик", "Email", "Состанок", "Демо", "Follow-up", "Друго"]
    for i, activity in enumerate(activities, 4):
        ws[f'E{i}'] = activity

    # Pipeline phases
    ws['G3'] = "Фази на pipeline:"
    ws['G3'].font = Font(bold=True)
    phases = ["Квалификација", "Демо", "Понуда", "Преговори", "Затворено-Добиено", "Затворено-Изгубено"]
    for i, phase in enumerate(phases, 4):
        ws[f'G{i}'] = phase

    # Products
    ws['I3'] = "Производи/Планови:"
    ws['I3'].font = Font(bold=True)
    products = ["Основен план", "Про план", "Премиум план", "Enterprise"]
    for i, product in enumerate(products, 4):
        ws[f'I{i}'] = product

    # Cities
    ws['A15'] = "Градови:"
    ws['A15'].font = Font(bold=True)
    cities = ["Скопје", "Битола", "Куманово", "Прилеп", "Тетово", "Охрид", "Велес", "Штип", "Гевгелија", "Кавадарци", "Струмица", "Кочани"]
    for i, city in enumerate(cities, 16):
        ws[f'A{i}'] = city

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['I'].width = 20

    return ws

def create_instructions_sheet(wb):
    """Create instructions sheet"""
    ws = wb.create_sheet("ℹ️ Упатство", 7)

    ws['A1'] = "УПАТСТВО ЗА КОРИСТЕЊЕ НА SMART CRM"
    ws['A1'].font = Font(size=18, bold=True, color="4472C4")
    ws.merge_cells('A1:D1')

    instructions = [
        "",
        "📊 DASHBOARD",
        "• Прегледот автоматски се ажурира според податоците во другите листови",
        "• Сите формули се автоматски - не ги менувајте",
        "",
        "🏢 КОМПАНИИ",
        "• Додавајте нови компании на крајот од листата",
        "• Користете dropdown менија за Статус и Приоритет",
        "• Зелено = Добиен, Црвено = Изгубен, Жолто = Заинтересиран",
        "• Филтрирајте по град, статус или извор",
        "",
        "👥 КОНТАКТИ",
        "• Поврзете контакти со компании преку име на компанија",
        "• Означете примарен контакт за секоја компанија",
        "",
        "📅 АКТИВНОСТИ",
        "• Закажувајте повици, emails и состаноци",
        "• Црвено = закаснета задача",
        "• Ажурирајте статус по завршување",
        "",
        "💰 PIPELINE",
        "• Следете deals и нивната вредност",
        "• Веројатност помага за forecasting",
        "",
        "📧 EMAIL TEMPLATES",
        "• Готови шаблони за брзо праќање",
        "• Персонализирајте ги пред испраќање",
        "",
        "⚙️ ПОСТАВКИ",
        "• Тука се дефинирани сите dropdown опции",
        "• Додадете нови опции ако ви требаат",
        "",
        "💡 СОВЕТИ",
        "• Секој ден ажурирајте го статусот на активностите",
        "• Користете филтри за да најдете специфични записи",
        "• Редовно backup-ирајте го документот",
        "• Споделете преку Google Sheets за тимска работа",
    ]

    for i, line in enumerate(instructions, 3):
        ws[f'A{i}'] = line
        if line.startswith("📊") or line.startswith("🏢") or line.startswith("👥") or \
           line.startswith("📅") or line.startswith("💰") or line.startswith("📧") or \
           line.startswith("⚙️") or line.startswith("💡"):
            ws[f'A{i}'].font = Font(bold=True, size=12, color="4472C4")

    ws.column_dimensions['A'].width = 80

    return ws

def main():
    print("🚀 Creating Facturino Smart CRM...")

    # Load leads
    print("📥 Loading leads from CSV files...")
    leads = load_leads()
    print(f"   Found {len(leads)} leads")

    # Create workbook
    print("📝 Creating workbook...")
    wb = create_workbook()

    # Create sheets
    print("   Creating Dashboard...")
    create_dashboard(wb)

    print("   Creating Companies sheet...")
    create_companies_sheet(wb, leads)

    print("   Creating Contacts sheet...")
    create_contacts_sheet(wb)

    print("   Creating Activities sheet...")
    create_activities_sheet(wb)

    print("   Creating Pipeline sheet...")
    create_pipeline_sheet(wb)

    print("   Creating Email Templates...")
    create_email_templates_sheet(wb)

    print("   Creating Settings...")
    create_settings_sheet(wb)

    print("   Creating Instructions...")
    create_instructions_sheet(wb)

    # Save workbook
    print(f"💾 Saving to {OUTPUT_FILE}...")
    wb.save(OUTPUT_FILE)

    print(f"""
✅ Smart CRM created successfully!

📁 File: {OUTPUT_FILE}

📋 Sheets included:
   1. 📊 Dashboard - Overview with live stats
   2. Компании - {len(leads)} companies with status tracking
   3. Контакти - Contact management
   4. Активности - Activity/task tracking
   5. Pipeline - Sales deals tracking
   6. Email Templates - Ready-to-use templates
   7. ⚙️ Поставки - Dropdown options
   8. ℹ️ Упатство - How to use guide

🔗 To use with Google Sheets:
   1. Go to drive.google.com
   2. Click "New" → "File upload"
   3. Select the .xlsx file
   4. Open the file and click "Open with Google Sheets"
""")

if __name__ == "__main__":
    main()
